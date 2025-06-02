from typing import List, Optional
from openpyxl import load_workbook
from pydantic import BaseModel
from app.main.service.import_classes import ExcelImporter
from app.main.models import Classroom, Discipline, Group, Specialty, Student, Teacher
from app.main.services import get_groups, get_specialties, get_teachers, send_entity
import io


def import_excel(api: str, entity: str, file_stream):
    results = ImportResults(success=0, errors=[], total=0)
    wb = load_workbook(filename=file_stream)
    data_to_send = []
    if wb.active:
        ws = wb.active

        headers = [cell.value for cell in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            results.total += 1
            row_data = dict(zip(headers, row))

            data = {}
            for header in row_data.keys():
                h = entity_map[entity][header]
                data[h] = row_data[header]

            data_to_send.append(data)

    print(data_to_send)

    return send_entity(
        f'{api}/{entity}',
        data_to_send,
        "Специальности импортированны"
    )
