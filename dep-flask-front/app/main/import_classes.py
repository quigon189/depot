import io
from typing import Dict, List, Optional, Type, TypeVar

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import BaseModel
from requests import get


T = TypeVar('T', bound=BaseModel)


class ExcelImporter:
    """
    Общий клас экспорта из Excel файла
    Так же нужен для создания шаблонов Excel файлов
    """

    def __init__(self, model: Type[T], template_properties: List[str], dependensies: Optional[Dict] = None):
        self.model = model
        self.dependensies = dependensies or {}
        self.template_properties = template_properties
        schema = model.model_json_schema()
        m = schema['$ref'].split('/')[-1]
        self.schema = schema['$defs'][m]

    def generate_template(self) -> io.BytesIO:
        wb = Workbook()
        ws = wb.active or wb.worksheets[0]

        ws.title = self.schema.get('title', 'Data')

        fields = self._get_fields()

        headers = []
        for field_info in fields:
            if field_info['name'] in self.template_properties:
                headers.append(field_info['header'])

        ws.append(headers)

        if self.dependensies:
            list_sheet = wb.create_sheet('Списки')
            i = 1
            for d_name, dependence in self.dependensies.items():
                letter = get_column_letter(i)
                i += 1
                list_sheet[f'{letter}1'] = d_name
                for idx, value in enumerate(dependence, start=2):
                    cell = f'{letter}{idx}'
                    list_sheet[cell] = value

                dv = DataValidation(
                    type='list',
                    formula1=f'=Списки!${letter}$2:${letter}${len(dependence)+1}',
                    allow_blank=True
                )
                letter_dep = get_column_letter(
                    self.template_properties.index(d_name)+1)
                dv.add(f'{letter_dep}2:{letter_dep}1048576')
                ws.add_data_validation(dv)

        for col in ws.columns:
            max_lenght = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_lenght:
                        max_lenght = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = (max_lenght + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _get_fields(self) -> List[Dict]:
        fields = []

        for field_name, field_info in self.schema['properties'].items():
            fields.append({
                'name': field_name,
                'header': field_info.get('title', field_name.replace('_', ' ').upper()),
                'type': field_info.get('type', 'string'),
                'required': field_name in self.schema.get('required', [])
            })

        return fields
