import io
from typing import Dict, List, Optional, Tuple, Type, TypeVar

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, ValidationError

from app.main.models import Classroom, Discipline, Group, Specialty, Student, Teacher
from app.main.service.view_services import get_groups, get_specialties, get_teachers, send_entity


T = TypeVar('T', bound=BaseModel)


class TemplateProperties:
    """
    Содержит все свойства шаблона, такие как:
        - headers
        - hidden_headers
        _ reference_tables
    """

    def __init__(self,
                 headers: List[str],
                 references: Optional[Dict[tuple, List[tuple]]] = None):
        self.headers = headers
        self.references = references or {}
        if references:
            self.hidden_headers = [h for (_,h) in references.keys()]
        else:
            self.hidden_headers = []
        self.data_list = "Данные"
        self.ref_list = "Ссылки"

    @property
    def data_headers(self) -> List[str]:
        return self.hidden_headers + self.headers

    @property
    def ref_headers(self) -> List[str]:
        rh = []
        for (r1,r2) in self.references.keys():
            rh.append(str(r1))
            rh.append(str(r2))
        return rh

    def ref_values(self, ref: Tuple[str,str]) -> List[Tuple[str,str]]:
        return [r for r in self.references[ref]]



class ExcelImporter:
    """
    Общий класс экспорта из Excel файла
    Так же нужен для создания шаблонов Excel файлов
    """

    def __init__(self, model: Type[T], props: TemplateProperties):
        self.model = model
        self.props = props
        schema = model.model_json_schema()
        m = schema['$ref'].split('/')[-1]
        self.schema = schema['$defs'][m]
        self.max_records = 1000

    def generate_template(self) -> io.BytesIO:
        wb = Workbook()
        ws = wb.active or wb.worksheets[0]

        ws.title = self.props.data_list

        fields = self._get_fields()

        for col_num, header in enumerate(self.props.data_headers, start=1):
            ws.cell(row=1, column=col_num, value=fields[header]['header'])
            if fields[header]['type'] == 'string':
                for row in range(2, self.max_records+2):
                    ws.cell(row=row, column=col_num).number_format = "@"
            if header in self.props.hidden_headers:
                col_letter = ws.cell(row=1, column=col_num).column_letter
                ws.column_dimensions[col_letter].hidden = True

        if self.props.ref_headers:
            rs = wb.create_sheet(self.props.ref_list)
            ref_headers = [fields[h]['header'] for h in self.props.ref_headers]
            rs.append(ref_headers)
            refs = self.props.references
            for (rh1, rh2), data in refs.items():
                idx1 = self.props.ref_headers.index(rh1) + 1
                idx2 = self.props.ref_headers.index(rh2) + 1

                let1 = get_column_letter(idx1)
                let2 = get_column_letter(idx2)

                for row, (rd1, rd2) in enumerate(data, start=2):
                    rs[f'{let1}{row}'] = rd1
                    rs[f'{let2}{row}'] = rd2

                dv = DataValidation(
                        type='list',
                        formula1=f'={self.props.ref_list}!${let1}$2:${let1}${len(data)+1}',
                        showErrorMessage=True
                        )

                if rh1 in self.props.data_headers:
                    dv_idx = self.props.data_headers.index(rh1)+1
                    dv_let = get_column_letter(dv_idx)

                    dv.add(f'{dv_let}2:{dv_let}{self.max_records+1}')
                    ws.add_data_validation(dv)

                    if rh2 in self.props.data_headers:
                        f_idx = self.props.data_headers.index(rh2)+1
                        f_let = get_column_letter(f_idx)

                        for row in range(2, self.max_records+2):
                            ws[f'{f_let}{row}'].value = f'=IFERROR(VLOOKUP({dv_let}{row}, {self.props.ref_list}!${let1}$2:${let2}${len(data)+1}, 2, FALSE), "")'

        for col in ws.columns:
            column = col[0].column_letter
            adjusted_width = 22 * 1.2
            ws.column_dimensions[column].width = adjusted_width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def import_data(self, file_stream: io.BufferedReader, skip_rows: int = 1) -> List[T]:
        try:
            wb = load_workbook(filename=file_stream, data_only=True)
            ws = wb.worksheets[0]

            header_row = list(ws.iter_rows(
                min_row=1, max_row=1, values_only=True))[0]
            headers = [str(cell).strip() for cell in header_row]

            if not headers:
                raise ValueError(
                    "Отсутствуют заголовки в первой строке шаблона")

            missing_headers = []
            fields = self._get_fields()
            for h in self.props.headers:
                if not fields[h]['header'] in headers:
                    missing_headers.append(fields[h]['header'])

            if missing_headers:
                raise ValueError(
                    "В шаблоне отсутствуют следующие столбцы: " + ", ".join(missing_headers)
                )

            models = []
            errors = []

            ref_headers = [h for (h,_) in self.props.references.keys()]
            for row_idx, row in enumerate(ws.iter_rows(min_row=skip_rows+1, values_only=True), start=skip_rows+1):
                if not any(row):
                    continue

                row_dict = self._row_to_dict(row, headers)

                for header in ref_headers:
                    if header in row_dict.keys():
                        del row_dict[header]

                try:
                    model = self.model(**row_dict)
                    models.append(model)
                except ValidationError as e:
                    errors.append({
                        'row': row_idx,
                        'errors': e.errors(),
                        'data': row_dict
                    })

            if errors:
                raise ValueError(
                    'Ошибка валидации данных:\n',
                    ''.join([str(e) for e in errors])
                )

            return models

        except Exception as e:
            raise ValueError(str(e).strip())

    def _get_fields(self) -> Dict[str,Dict]:
        fields = {}

        for field_name, field_info in self.schema['properties'].items():
            fields[field_name] = {
                'header': field_info.get('title', field_name),
                'type': field_info.get('type', 'string'),
                'required': field_name in self.schema.get('required', [])
            }

        return fields

    def _row_to_dict(self, row: tuple, headers: List[str]) -> Dict:
        row_dict = {}

        fields = self._get_fields()

        for header, value in zip(headers, row):
            for name, field in fields.items():
                if field['header'] == header:
                    row_dict[name] = self._parse_value(value, field)
                    break

        return row_dict

    def _get_list_data(self, ls: Worksheet) -> List[Dict]:
        headers = [str(cell.value).strip() for cell in ls[1]]

        data = []
        for row in ls.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for header, cell_value in zip(headers, row):
                row_dict[header] = cell_value

            data.append(row_dict)

        return data

    def _parse_value(self, value: str, field: Dict):
        try:
            if field['type'] == 'integer':
                return int(value)
            elif field['type'] == 'number':
                return float(value)
            else:
                return str(value).strip() if value else None
        except (ValueError, TypeError):
            return value


class SpecialtyImporter(ExcelImporter):
    def __init__(self, api: str):
        super().__init__(
            model=Specialty,
            props=TemplateProperties(headers=['code', 'name', 'short_name'])
        )


class GroupImporter(ExcelImporter):
    def __init__(self, api: str):
        specialties = get_specialties(api).items
        teachers = get_teachers(api).items
        super().__init__(
            model=Group,
            props=TemplateProperties(
                headers=['number', 'year_formed', 'specialty', 'class_teacher'],
                references={
                    ('specialty','spec_id') : [(s.code, s.id) for s in specialties],
                    ('class_teacher','class_teacher_id') : [(t.name, t.id) for t in teachers]
                    }
                )
        )


class StudentImporter(ExcelImporter):
    def __init__(self, api: str):
        groups = get_groups(api).items
        super().__init__(
            model=Student,
            props=TemplateProperties(
                headers=['last_name', 'first_name', 'middle_name', 'group', 'birth_date', 'phone'],
                references={
                    ('group','group_id'): [(g.name, g.id) for g in groups]
                    }
                )
        )


class TeacherImporter(ExcelImporter):
    def __init__(self, api: str):
        super().__init__(
            model=Teacher,
            props=TemplateProperties(
                headers=['last_name', 'first_name', 'middle_name', 'birth_date', 'phone']
                )
        )


class DisciplineImporter(ExcelImporter):
    def __init__(self, api: str):
        groups = get_groups(api).items
        super().__init__(
            model=Discipline,
            props=TemplateProperties(
                headers=['code', 'name', 'semester', 'hours', 'group'],
                references={
                    ('group','group_id'): [(g.name, g.id) for g in groups]
                    }
                )
        )


class ClassroomImporter(ExcelImporter):
    def __init__(self, api: str):
        teachers = get_teachers(api).items
        super().__init__(
            model=Classroom,
            props=TemplateProperties(
                headers=['number', 'name', 'type', 'capacity', 'equipment', 'teacher'],
                references={
                    ('teacher', 'teacher_id'): [(t.name, t.id) for t in teachers],
                    }
                )
        )


importer_switch = {
    'specialties': SpecialtyImporter,
    'groups': GroupImporter,
    'students': StudentImporter,
    'teachers': TeacherImporter,
    'disciplines': DisciplineImporter,
    'classes': ClassroomImporter
}


def generate_template(entity: str, api: str) -> Optional[io.BytesIO]:
    if entity in importer_switch.keys():
        return importer_switch[entity](api).generate_template()


def import_from_template(api: str, entity: str, file_stream: io.BufferedReader) -> Tuple[str, str]:
    if entity in importer_switch.keys():
        try:
            models = importer_switch[entity](api).import_data(file_stream)
            return send_entity(
                f'{api}/{entity}',
                [m.to_dict() for m in models],
                f'Количество успешно импортированных записей {len(models)}'
            )
        except ValueError as e:
            return str(e), 'warning'

    return 'Недопустимое значение сущности', 'danger'
