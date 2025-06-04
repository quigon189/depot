import io
import json
from typing import Dict, List, Optional, Tuple, Type, TypeVar

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, ValidationError

from app.main.models import Classroom, Discipline, Group, Specialty, Student, Teacher
from app.main.services import get_groups, get_specialties, get_teachers, send_entity


T = TypeVar('T', bound=BaseModel)


class ExcelImporter:
    """
    Общий клас экспорта из Excel файла
    Так же нужен для создания шаблонов Excel файлов
    """

    def __init__(self, model: Type[T], template_properties: List[str], dependensies: Optional[List[Dict]] = None):
        self.model = model
        self.dependensies = dependensies or {}
        self.template_properties = template_properties
        schema = model.model_json_schema()
        m = schema['$ref'].split('/')[-1]
        self.schema = schema['$defs'][m]

    def generate_template(self) -> io.BytesIO:
        wb = Workbook()
        ws = wb.active or wb.worksheets[0]

        ws.title = self.schema.get('title', 'Data')

        fields = self._get_fields()

        headers = []
        for field_info in fields:
            hidden = not field_info['name'] in self.template_properties
            headers.append({
                'header': field_info['header'],
                'hidden': hidden
                })

        headers = sorted(headers, key=lambda x: not x['hidden'])
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_num, value=header['header'])
            if header['hidden']:
                col_letter = ws.cell(row=1, column=col_num).column_letter
                ws.column_dimensions[col_letter].hidden = True

        if self.dependensies:
            list_sheet = wb.create_sheet('Списки')
            i = 1
            for dependence in self.dependensies:
                letter_value = get_column_letter(i)
                letter_id = get_column_letter(i+1)
                v_header, id_header = dependence['headers']
                list_sheet[f'{letter_value}1'] = v_header
                list_sheet[f'{letter_id}1'] = id_header
                for idx, (value, id) in enumerate(dependence['values'], start=2):
                    list_sheet[f'{letter_value}{idx}'] = value
                    list_sheet[f'{letter_id}{idx}'] = id

                if v_header in self.template_properties:

                    dv = DataValidation(
                        type='list',
                        formula1=f'=Списки!${letter_value}$2:${letter_value}${len(dependence["values"])+1}',
                        showErrorMessage=True
                    )
                    for idx, header in enumerate(headers, start=1):
                        for field_info in fields:
                            if field_info['name'] == v_header and field_info['header'] == header['header']:
                                letter_dep = get_column_letter(idx)

                                dv.add(f'{letter_dep}2:{letter_dep}1048576')
                                ws.add_data_validation(dv)

                                for idx2, header2 in enumerate(headers, start=1):
                                    for field_info2 in fields:
                                        if field_info2['name'] == id_header and field_info2['header'] == header2['header']:
                                            for row in range(2, 10000):
                                                ws.cell(
                                                        row=row,
                                                        column=idx2,
                                                        value=f'=ЕСЛИОШИБКА(ВПР({letter_dep}{row};Списки!{letter_value}2:{letter_id}{len(dependence["values"])};2;ЛОЖЬ);"0")'
                                                        )
                i += 2

        for col in ws.columns:
            column = col[0].column_letter
            adjusted_width = 22 * 1.2
            ws.column_dimensions[column].width = adjusted_width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def import_data(self, file_stream: io.BufferedReader, skip_rows: int = 1) -> List[T]:
        try:
            wb = load_workbook(filename=file_stream)
            ws = wb.worksheets[0]

            header_row = list(ws.iter_rows(
                min_row=1, max_row=1, values_only=True))[0]
            headers = [str(cell).strip() for cell in header_row]

            if not headers:
                raise ValueError(
                    "Отсутствуют заголовки в первой строке шаблона")

            missing_fields = []
            fields = self._get_fields()
            for propertie in self.template_properties:
                for field in fields:
                    if field['name'] == propertie and not field['header'] in headers:
                        missing_fields.append(field)
                        break

            if missing_fields:
                raise ValueError(
                    "В шаблоне отсутствуют следующие столбцы: ",
                    "".join([f['header'] for f in missing_fields])
                )

            models = []
            errors = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=skip_rows+1, values_only=True), start=skip_rows+1):
                row_dict = self._row_to_dict(row, headers)

                try:
                    model = self.model(**row_dict)
                    models.append(model)
                except ValidationError as e:
                    errors.append({
                        'row': row_idx,
                        'errors': e.errors(),
                        'data': row_dict
                    })

            if errors:
                raise ValueError(
                    'Ошибка валидации данных:\n',
                    ''.join([str(e) for e in errors])
                )

            return models

        except Exception as e:
            raise ValueError(str(e).strip())

    def _get_fields(self) -> List[Dict]:
        fields = []

        for field_name, field_info in self.schema['properties'].items():
            fields.append({
                'name': field_name,
                'header': field_info.get('title', field_name),
                'type': field_info.get('type', 'string'),
                'required': field_name in self.schema.get('required', [])
            })

        return fields

    def _row_to_dict(self, row: tuple, headers: List[str]) -> Dict:
        row_dict = {}

        fields = self._get_fields()

        for header, value in zip(headers, row):
            for field in fields:
                if field['header'] == header:
                    row_dict[field['name']] = self._parse_value(value, field)
                    break

        return row_dict

    def _get_list_data(self, ls: Worksheet) -> List[Dict]:
        headers = [str(cell.value).strip() for cell in ls[1]]

        data = []
        for row in ls.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for header, cell_value in zip(headers, row):
                row_dict[header] = cell_value

            data.append(row_dict)

        return data

    def _parse_value(self, value: str, field: Dict):
        try:
            if field['type'] == 'integer':
                return int(value)
            elif field['type'] == 'number':
                return float(value)
            else:
                return str(value).strip() if value else None
        except (ValueError, TypeError):
            return value


class SpecialtyImporter(ExcelImporter):
    def __init__(self, api: str):
        super().__init__(
            model=Specialty,
            template_properties=['code', 'name', 'short_name']
        )


class GroupImporter(ExcelImporter):
    def __init__(self, api: str):
        specialties = get_specialties(api).items
        teachers = get_teachers(api).items
        super().__init__(
            model=Group,
            template_properties=['number', 'year_formed',
                                 'specialty', 'class_teacher'],
            dependensies=[
                {
                    'headers': ('specialty', 'spec_id'),
                    'values': [(s.code, s.id) for s in specialties]
                    },
                {
                    'headers': ('class_teacher', 'class_teacher_id'),
                    'values': [(t.name, t.id) for t in teachers]
                    }
                ]
        )


class StudentImporter(ExcelImporter):
    def __init__(self, api: str):
        groups = get_groups(api).items
        super().__init__(
            model=Student,
            template_properties=['last_name', 'first_name', 'middle_name',
                                 'group', 'birth_date', 'phone'],
            dependensies={
                'group': [g.name for g in groups],
                'group_id': [g.id for g in groups],
            }
        )


class TeacherImporter(ExcelImporter):
    def __init__(self, api: str):
        super().__init__(
            model=Teacher,
            template_properties=['last_name', 'first_name', 'middle_name',
                                 'birth_date', 'phone'],
        )


class DisciplineImporter(ExcelImporter):
    def __init__(self, api: str):
        groups = get_groups(api).items
        super().__init__(
            model=Discipline,
            template_properties=['code', 'name', 'semester', 'hours', 'group'],
            dependensies={
                'group': [g.name for g in groups],
                'group_id': [g.id for g in groups],
            }
        )


class ClassroomImporter(ExcelImporter):
    def __init__(self, api: str):
        teachers = get_teachers(api).items
        super().__init__(
            model=Classroom,
            template_properties=['number', 'name', 'type',
                                 'capacity', 'equipment', 'teacher'],
            dependensies={
                'type': ['Кабинет', 'Лаборатория', 'Полигон'],
                'teacher': [t.name for t in teachers],
                'teacher_id': [t.id for t in teachers],
            }
        )


importer_switch = {
    'specialties': SpecialtyImporter,
    'groups': GroupImporter,
    'students': StudentImporter,
    'teachers': TeacherImporter,
    'disciplines': DisciplineImporter,
    'classes': ClassroomImporter
}


def generate_template(entity: str, api: str) -> Optional[io.BytesIO]:
    if entity in importer_switch.keys():
        return importer_switch[entity](api).generate_template()


def import_from_template(api: str, entity: str, file_stream: io.BufferedReader) -> Tuple[str, str]:
    if entity in importer_switch.keys():
        try:
            models = importer_switch[entity](api).import_data(file_stream)
            return send_entity(
                f'{api}/{entity}',
                [m.to_dict() for m in models],
                f'Количество успешно импортированных записей {len(models)}'
            )
        except ValueError as e:
            return str(e), 'warning'

    return 'Недопустимое значение сущности', 'danger'
